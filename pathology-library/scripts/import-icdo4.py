#!/usr/bin/env python3
"""Build the browser ICD-O-4 catalogue from the official IARC Excel tables."""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ICDO_SOURCE = "https://tumourclassification.iarc.who.int/icd-o-4/index.php"
ICDO_RELEASE = "https://www.iarc.who.int/news-events/international-classification-of-diseases-for-oncology-fourth-edition-icd-o-4-final-tables-now-available/"


def clean(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def unique(values: Iterable[Any]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        item = clean(value)
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def rows_as_dicts(sheet) -> Iterable[dict[str, Any]]:
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if any(value is not None for value in values):
            yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}


def parse_morphology(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in rows_as_dicts(workbook["Morphology"]):
        code = clean(row.get("ICDO4")).upper()
        level = clean(row.get("Level"))
        # Numeric levels and code ranges are hierarchy labels, not billable
        # morphology entries. Keep only rows carrying a term relationship.
        if level.isdigit():
            continue
        if not code:
            continue
        entry = groups.setdefault(code, {
            "code": code,
            "preferred": "",
            "terms": [],
            "codeReferences": [],
            "seeAlso": [],
            "notes": [],
            "includes": [],
            "excludes": [],
            "otherText": [],
        })
        term = clean(row.get("Term"))
        if term:
            entry["terms"].append({"type": level or "Related", "term": term})
            if level.lower() == "preferred" and not entry["preferred"]:
                entry["preferred"] = term
        for source, target in (
            ("Code reference", "codeReferences"),
            ("See also", "seeAlso"),
            ("See note", "notes"),
            ("Includes", "includes"),
            ("Excludes", "excludes"),
            ("Other text", "otherText"),
        ):
            value = clean(row.get(source))
            if value:
                entry[target].append(value)

    code_pattern = re.compile(r"^(\d{4})([0-9A-Z])/([0-9A-Z])$")
    output: list[dict[str, Any]] = []
    for entry in groups.values():
        if not entry["preferred"] and entry["terms"]:
            entry["preferred"] = entry["terms"][0]["term"]
        match = code_pattern.match(entry["code"])
        entry["legacyCode"] = f"{match.group(1)}/{match.group(3)}" if match else ""
        entry["extension"] = match.group(2) if match else ""
        entry["behaviour"] = match.group(3) if match else ""
        for key in ("codeReferences", "seeAlso", "notes", "includes", "excludes", "otherText"):
            entry[key] = unique(entry[key])
        output.append(entry)
    return output


def parse_topography_sheet(sheet, optional: bool) -> list[dict[str, Any]]:
    groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    headings: dict[int, dict[str, str]] = {}
    for row in rows_as_dicts(sheet):
        code = clean(row.get("ICDO4")).upper()
        level = clean(row.get("Level"))
        term = clean(row.get("Term"))
        if level.isdigit():
            depth = int(level)
            headings[depth] = {"code": code, "term": term}
            headings = {key: value for key, value in headings.items() if key <= depth}
            continue
        if not code or not term:
            continue
        entry = groups.setdefault(code, {
            "code": code,
            "preferred": "",
            "terms": [],
            "hierarchy": [],
            "optional": optional,
            "notes": [],
            "codeReferences": [],
            "seeAlso": [],
            "seeNotes": [],
            "includes": [],
            "excludes": [],
            "otherText": [],
        })
        entry["terms"].append({"type": level or "Related", "term": term})
        if level.lower() == "preferred" and not entry["preferred"]:
            entry["preferred"] = term
            entry["hierarchy"] = [value for _, value in sorted(headings.items()) if value.get("term")]
        for source, target in (
            ("Note", "notes"),
            ("code_reference", "codeReferences"),
            ("Code reference", "codeReferences"),
            ("See also", "seeAlso"),
            ("See note", "seeNotes"),
            ("Includes", "includes"),
            ("Excludes", "excludes"),
            ("Other text", "otherText"),
        ):
            value = clean(row.get(source))
            if value:
                entry[target].append(value)

    output: list[dict[str, Any]] = []
    for entry in groups.values():
        if not entry["preferred"] and entry["terms"]:
            entry["preferred"] = entry["terms"][0]["term"]
        for key in ("notes", "codeReferences", "seeAlso", "seeNotes", "includes", "excludes", "otherText"):
            entry[key] = unique(entry[key])
        output.append(entry)
    return output


def load_annex_rows(path: Path) -> dict[str, list[dict[str, str]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}
    for sheet in workbook.worksheets:
        headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=3, values_only=True):
            item = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers) and clean(value)}
            if item:
                rows.append(item)
        result[sheet.title] = rows
    return result


def first_code(row: dict[str, str], candidates: tuple[str, ...]) -> str:
    for key in candidates:
        if row.get(key):
            return row[key].upper()
    return ""


def morphology_changes(annexes: dict[str, list[dict[str, str]]]) -> tuple[dict[str, list[dict[str, str]]], list[dict[str, str]]]:
    changes: dict[str, list[dict[str, str]]] = defaultdict(list)
    deleted: list[dict[str, str]] = []
    labels = {
        "New morphology codes (4 digits)": "new-four-digit",
        "New morphology codes (5 digits)": "new-five-digit",
        "Morphology code changes": "code-change",
        "Behaviour code changes": "behaviour-change",
        "New morphology terms": "new-term",
        "Morphology term changes": "term-change",
    }
    for sheet_name, rows in annexes.items():
        if sheet_name == "Deleted morphology codes":
            deleted.extend(rows)
            continue
        change_type = labels.get(sheet_name)
        if not change_type:
            continue
        for row in rows:
            code = first_code(row, ("ICDO4",))
            if code:
                changes[code].append({"type": change_type, **row})
    return changes, deleted


def topography_changes(annexes: dict[str, list[dict[str, str]]]) -> dict[str, list[dict[str, str]]]:
    changes: dict[str, list[dict[str, str]]] = defaultdict(list)
    labels = {
        "Topography new codes": "new-code",
        "Topography code changes": "code-change",
        "New topography terms": "new-term",
        "Topography term changes": "term-change",
        "New optional topography codes": "new-optional-code",
        "New optional topography terms": "new-optional-term",
        "Optional topography term change": "optional-term-change",
    }
    for sheet_name, rows in annexes.items():
        change_type = labels.get(sheet_name)
        if not change_type:
            continue
        for row in rows:
            code = first_code(row, ("ICDO4", "ICDO4 Code", "ICDO4 Optional code", "ICDO4 optional"))
            if code:
                changes[code].append({"type": change_type, **row})
            optional_code = first_code(row, ("Optional code", "ICDO4 Optional code", "ICDO4 optional"))
            if optional_code and optional_code != code:
                changes[optional_code].append({"type": change_type, **row})
    return changes


def compact_changes(items: list[dict[str, str]]) -> list[dict[str, str]]:
    keys = (
        "type", "Term", "ICDO3.2", "Term 3.2", "Change type", "Reason",
        "ICDO4 Code", "ICDO4 Term", "ICDO4 Optional code", "ICDO4 Optional term",
    )
    return [{key: item[key] for key in keys if item.get(key)} for item in items]


def attach_changes(entries: list[dict[str, Any]], changes: dict[str, list[dict[str, str]]]) -> None:
    for entry in entries:
        entry["changes"] = compact_changes(changes.get(entry["code"], []))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--catalog", required=True, type=Path)
    parser.add_argument("--morphology-annexes", required=True, type=Path)
    parser.add_argument("--topography-annexes", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    for path in (args.catalog, args.morphology_annexes, args.topography_annexes):
        if not path.exists():
            raise FileNotFoundError(path)

    morphology = parse_morphology(args.catalog)
    catalog_workbook = openpyxl.load_workbook(args.catalog, read_only=True, data_only=True)
    topography = parse_topography_sheet(catalog_workbook["Topography"], optional=False)
    morphology_annexes = load_annex_rows(args.morphology_annexes)
    topography_annexes = load_annex_rows(args.topography_annexes)
    optional_workbook = openpyxl.load_workbook(args.topography_annexes, read_only=True, data_only=True)
    optional_topography = parse_topography_sheet(optional_workbook["Optional topography 4-digit"], optional=True)

    morph_change_map, deleted_morphology = morphology_changes(morphology_annexes)
    topo_change_map = topography_changes(topography_annexes)
    attach_changes(morphology, morph_change_map)
    attach_changes(topography, topo_change_map)
    attach_changes(optional_topography, topo_change_map)

    payload = {
        "meta": {
            "edition": "ICD-O-4",
            "title": "International Classification of Diseases for Oncology, Fourth Edition",
            "releaseDate": "2026-07-20",
            "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "sourceUrl": ICDO_SOURCE,
            "releaseUrl": ICDO_RELEASE,
            "sourceFiles": [args.catalog.name, args.morphology_annexes.name, args.topography_annexes.name],
            "morphologyCodeCount": len(morphology),
            "morphologyTermCount": sum(len(item["terms"]) for item in morphology),
            "topographyCodeCount": len(topography),
            "topographyTermCount": sum(len(item["terms"]) for item in topography),
            "optionalTopographyCodeCount": len(optional_topography),
            "optionalTopographyTermCount": sum(len(item["terms"]) for item in optional_topography),
            "deletedMorphologyCodeCount": len(deleted_morphology),
        },
        "behaviours": {
            "0": {"vi": "Lành tính", "en": "Benign"},
            "1": {"vi": "Không xác định lành tính hay ác tính", "en": "Uncertain whether benign or malignant"},
            "2": {"vi": "Tại chỗ", "en": "In situ"},
            "3": {"vi": "Ác tính, vị trí nguyên phát", "en": "Malignant, primary site"},
            "6": {"vi": "Ác tính, vị trí di căn", "en": "Malignant, metastatic site"},
            "9": {"vi": "Ác tính, chưa xác định nguyên phát hay di căn", "en": "Malignant, uncertain whether primary or metastatic"},
        },
        "morphology": morphology,
        "topography": topography,
        "optionalTopography": optional_topography,
        "deletedMorphologyCodes": deleted_morphology,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps(payload["meta"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
